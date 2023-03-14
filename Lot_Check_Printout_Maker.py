import os
from openpyxl.styles import Border, Side
import openpyxl as xl
from tkinter import filedialog
import tkinter as tk

# Create the main window
root = tk.Tk()

# Hide the main window since we don't need it
root.withdraw()

# Ask the user to select an Excel file
file_path = filedialog.askopenfilename(
    title="Select an Excel file",
    filetypes=[("Excel files", "*.xlsx;*.xls")],
)

# Check if a file was selected
if not file_path:
    print("No file selected.")
    exit()

# Load the Excel file with openpyxl
workbook = xl.load_workbook(filename=file_path)

# Get the first sheet of the workbook
sheet = workbook.active

# Create a new worksheet
new_sheet = workbook.create_sheet(title="Check", index=0)

# Get the folder and file name from the original file path
folder_path, file_name = os.path.split(file_path)
file_base_name, file_extension = os.path.splitext(file_name)

# Create the new file path
new_file_name = file_base_name + "_Check" + file_extension
new_file_path = os.path.join(folder_path, new_file_name)

# Check if the new file name already exists, and rename it if necessary
if os.path.exists(new_file_path):
    partTwo = True
    old_file_name = file_base_name + "_Check_Old" + file_extension
    old_file_path = os.path.join(folder_path, old_file_name)
    if os.path.exists(old_file_path):
        os.remove(old_file_path)
    os.rename(new_file_path, old_file_path)
    print(f"Old file renamed: {old_file_path}")
else:
    partTwo = False

# Copy the first row from the original sheet to the new sheet
for col_num, cell in enumerate(sheet[1], start=1):
    new_sheet.cell(row=1, column=col_num, value=cell.value)

# Transpose the first row from the original sheet and write it as a column in the new sheet
for row_num, cell in enumerate(sheet[1], start=1):
    new_sheet.cell(row=row_num, column=1, value=cell.value)

new_sheet.delete_rows(1)
new_sheet.insert_rows(1)

# Add the text "Sampled:", "Logged:", and "Photo:" to the new sheet
new_sheet.cell(row=1, column=2, value="Sampled:")
new_sheet.cell(row=1, column=3, value="Logged:")
new_sheet.cell(row=1, column=4, value="Photo:")

# Save the new Excel file
workbook.save(new_file_path)

print(f"New file created: {new_file_path}")


####Part 2####
if partTwo == True:
    # Load the two input Excel files
    workbook1 = xl.load_workbook(new_file_path)
    workbook2 = xl.load_workbook(old_file_path)

    # Get the first worksheet of each input file
    worksheet1 = workbook1.active
    worksheet2 = workbook2.active

    # Create a new workbook to write the output to
    output_workbook = xl.Workbook()
    output_worksheet = output_workbook.active

    # Create a set to store the lot numbers that are found in both input files
    lot_numbers_in_both_files = set()

    
    # Loop over the rows of the first worksheet in the first input file
    current_row=2
    for row in worksheet1.iter_rows(min_row=2, max_col=1, values_only=True):
        # Get the lot number from the first column of the current row
        lot_number = row[0]

        # Check if the lot number is also in the first column of the second worksheet in the second input file
        if worksheet2.cell(column=1, row=current_row).value == lot_number:
            lot_numbers_in_both_files.add(lot_number)
        else:
            # The lot number is only in one of the input files, so write it to the output worksheet
            output_worksheet.append([lot_number])
        
        current_row += 1

    # set the border properties
    border = Border(left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin'))

    # iterate over all cells in the range (A1:C3)
    for row in output_worksheet.iter_rows(min_row=1, min_col=1, max_row=40, max_col=4):
        for cell in row:
            cell.border = border

    output_worksheet.insert_rows(1)
    # Add the text "Sampled:", "Logged:", and "Photo:" to the new sheet
    output_worksheet.cell(row=1, column=2, value="Sampled:")
    output_worksheet.cell(row=1, column=3, value="Logged:")
    output_worksheet.cell(row=1, column=4, value="Photo:")

    # Save the output workbook
    new_new_file_name = file_base_name + '_Print' + file_extension
    new_new_file_path = os.path.join(folder_path, new_new_file_name)

    if os.path.exists(new_new_file_path):
        os.remove(new_new_file_path)
    output_workbook.save(new_new_file_path)
